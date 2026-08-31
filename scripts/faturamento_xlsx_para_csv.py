#!/usr/bin/env python3
"""
Converte os Excels historicos de faturamento do TOTVS (RELATORIOS TOTVS\\Legacy)
para um CSV normalizado que o comando `legado:import-faturamento-arquivo` consome.

POR QUE PYTHON E NAO PHP: o PhpSpreadsheet mantem toda celula como objeto em
memoria (mesmo motivo do estouro de memory_limit ja documentado nos exports Excel
do CRM). O `2025 final` tem 285 MB e ~1,1 M de linhas -- openpyxl em read_only faz
streaming de verdade e roda com memoria constante.

VARIACOES DE LAYOUT QUE ESTE SCRIPT ABSORVE (todas reais, medidas em 30/08/2026):
  1. O cabecalho nao esta sempre na linha 1 -- esta na L2 no "2024 final" e na L3
     na aba "parte 1" do "2025 final". E localizado, nao assumido.
  2. A ORDEM das colunas muda: as abas "(2)" do 2025 final comecam com
     NOME_SUPERVISOR e nao com FILIAL. Por isso o mapeamento e SEMPRE por nome de
     coluna -- nunca por indice. Trocar isso por posicao fixa corrompe 1,1 M de
     linhas em silencio, porque os tipos ainda "cabem" nas colunas erradas.
  3. EMISSAO vem como texto "dd/mm/aaaa" nos arquivos .1/.2 e como datetime real
     nos arquivos "final".
  4. Numeros vem ora como float (openpyxl converteu) ora como texto com virgula
     decimal e separador de milhar.

Uso:
    python scripts/faturamento_xlsx_para_csv.py <entrada.xlsx> <saida.csv> [--abas "parte 1,parte 2"]

O CSV de saida tem cabecalho e usa as MESMAS colunas da tabela `faturamentos`.
"""

import argparse
import csv
import datetime as dt
import os
import re
import sys

from openpyxl import load_workbook

# Coluna do CSV de saida -> nome da coluna no Excel de origem.
# A chave e o destino no banco; o valor e o que procurar no cabecalho.
MAPA = {
    "filial": "FILIAL",
    "nota_fiscal": "NTA_FISCAL",
    "pedido": "PEDIDO",
    "data_emissao": "EMISSAO",
    "cod_cliente": "COD_CLI",
    "cnpj": "CNPJ",
    "cliente_nome": "CLIENTE",
    "cod_vendedor": "COD_VENDEDOR",
    "cod_produto": "COD_PROD",
    "produto_desc": "DES_PROD",
    "segmento": "SEGMENTO",
    "quantidade": "QUANT",
    "valor_unitario": "VLR_UNIT",
    "valor_total": "VLR_TOTAL",
}

# Sem estas o registro nao serve para nada no CRM.
OBRIGATORIAS = ("EMISSAO", "VLR_TOTAL")

SAIDA = list(MAPA.keys())


def norm_cabecalho(valor):
    """Normaliza para casar 'Estado      ' com 'ESTADO'."""
    return re.sub(r"\s+", " ", str(valor or "")).strip().upper()


def achar_cabecalho(ws, limite=8):
    """
    Localiza a linha de cabecalho e devolve (indice_da_linha, {NOME: indice_coluna}).

    Procura a primeira linha das `limite` iniciais que contenha EMISSAO e VLR_TOTAL --
    e nao simplesmente "a primeira linha com muitas celulas", porque a linha 1 dos
    arquivos "final" e o titulo do relatorio e tem uma celula so.
    """
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=limite, values_only=True), start=1):
        nomes = [norm_cabecalho(c) for c in row]
        if "EMISSAO" in nomes and "VLR_TOTAL" in nomes:
            indices = {}
            for pos, nome in enumerate(nomes):
                if nome and nome not in indices:
                    indices[nome] = pos
            return i, indices
    return None, None


def parse_data(valor):
    """Aceita datetime/date do openpyxl ou texto 'dd/mm/aaaa'. Devolve 'AAAA-MM-DD'."""
    if valor is None:
        return None
    if isinstance(valor, dt.datetime):
        return valor.date().isoformat()
    if isinstance(valor, dt.date):
        return valor.isoformat()
    texto = str(valor).strip()
    if not texto:
        return None
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})", texto)
    if m:
        d, mes, a = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return dt.date(a, mes, d).isoformat()
        except ValueError:
            return None
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", texto)
    if m:
        return m.group(0)
    return None


def parse_numero(valor):
    """
    Aceita float do openpyxl ou texto brasileiro ('1.234,56'). Preserva o sinal --
    valores negativos sao devolucoes e sao justamente o motivo dos arquivos
    'com negativos' existirem.
    """
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip()
    if not texto:
        return None
    negativo = texto.startswith("-")
    texto = texto.lstrip("+-").replace(".", "").replace(",", ".")
    texto = re.sub(r"[^\d.]", "", texto)
    if texto in ("", "."):
        return None
    try:
        n = float(texto)
    except ValueError:
        return None
    return -n if negativo else n


def texto(valor):
    if valor is None:
        return ""
    return re.sub(r"\s+", " ", str(valor)).strip()


def converter(entrada, saida, abas_escolhidas=None):
    wb = load_workbook(entrada, read_only=True, data_only=True)

    abas = abas_escolhidas or wb.sheetnames
    faltando = [a for a in abas if a not in wb.sheetnames]
    if faltando:
        print(f"ERRO: aba(s) inexistente(s): {faltando}", file=sys.stderr)
        print(f"      abas do arquivo: {wb.sheetnames}", file=sys.stderr)
        return 1

    total = escritas = sem_data = sem_valor = sem_vendedor = negativas = 0
    menor = maior = None

    with open(saida, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(SAIDA)

        for aba in abas:
            ws = wb[aba]
            hrow, indices = achar_cabecalho(ws)
            if hrow is None:
                print(f"  [aba {aba!r}] SEM cabecalho reconhecivel -- pulada", file=sys.stderr)
                continue

            ausentes = [orig for orig in MAPA.values() if orig not in indices]
            print(f"  [aba {aba!r}] cabecalho na L{hrow}; colunas ausentes: {ausentes or 'nenhuma'}")

            faltam_criticas = [c for c in OBRIGATORIAS if c not in indices]
            if faltam_criticas:
                print(f"  [aba {aba!r}] PULADA -- falta {faltam_criticas}", file=sys.stderr)
                continue

            for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
                total += 1

                def col(nome):
                    pos = indices.get(nome)
                    if pos is None or pos >= len(row):
                        return None
                    return row[pos]

                data = parse_data(col("EMISSAO"))
                if data is None:
                    sem_data += 1
                    continue

                valor_total = parse_numero(col("VLR_TOTAL"))
                if valor_total is None:
                    sem_valor += 1
                    continue
                if valor_total < 0:
                    negativas += 1

                cod_vendedor = texto(col("COD_VENDEDOR"))
                if cod_vendedor == "":
                    sem_vendedor += 1

                quantidade = parse_numero(col("QUANT"))
                valor_unitario = parse_numero(col("VLR_UNIT"))

                w.writerow([
                    texto(col("FILIAL")),
                    texto(col("NTA_FISCAL")),
                    texto(col("PEDIDO")),
                    data,
                    texto(col("COD_CLI")),
                    texto(col("CNPJ")),
                    texto(col("CLIENTE")),
                    cod_vendedor,
                    texto(col("COD_PROD")),
                    texto(col("DES_PROD")),
                    texto(col("SEGMENTO")),
                    "" if quantidade is None else quantidade,
                    "" if valor_unitario is None else valor_unitario,
                    valor_total,
                ])
                escritas += 1

                if menor is None or data < menor:
                    menor = data
                if maior is None or data > maior:
                    maior = data

                if escritas % 200000 == 0:
                    print(f"    {escritas} linhas escritas...")

    wb.close()

    print("")
    print(f"  lidas .............. {total}")
    print(f"  escritas ........... {escritas}")
    print(f"  sem data ........... {sem_data}")
    print(f"  sem valor .......... {sem_valor}")
    print(f"  sem cod_vendedor ... {sem_vendedor}   (entram, mas nao aparecem para nenhum vendedor)")
    print(f"  valor negativo ..... {negativas}   (devolucoes)")
    print(f"  periodo ............ {menor} .. {maior}")
    print(f"  saida .............. {saida} ({os.path.getsize(saida)/1048576:.1f} MB)")
    return 0


def main():
    p = argparse.ArgumentParser(description="Converte Excel de faturamento do TOTVS para CSV normalizado.")
    p.add_argument("entrada")
    p.add_argument("saida")
    p.add_argument("--abas", help="lista separada por virgula; por padrao todas as abas")
    args = p.parse_args()

    abas = [a.strip() for a in args.abas.split(",")] if args.abas else None
    print(f"Lendo {args.entrada}")
    sys.exit(converter(args.entrada, args.saida, abas))


if __name__ == "__main__":
    main()
